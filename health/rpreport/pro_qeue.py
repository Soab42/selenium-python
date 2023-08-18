from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from rpreport.credentials import usernames, passwords
import time
import json
import multiprocessing
from multiprocessing import Process, Queue
from openpyxl import Workbook

# Define the number of concurrent processes
MAX_CONCURRENT_PROCESSES = 1


def scrape_data(task_queue, result_queue):

    while True:
        try:
            username, password = task_queue.get_nowait()
        except multiprocessing.Queue.empty():
            break

        # Initialize Chrome WebDriver and scraping code
        # ... (same as your original scrape_data function)
        # Initialize Chrome WebDriver
        option = webdriver.ChromeOptions()
        option.add_experimental_option("detach", True)
        driver = webdriver.Chrome(options=option)

        # Set the desired IP address as the Host header
        ip_address = "103.139.165.110:8080"
        headers = {"Host": ip_address}

        # Navigate to the login page
        driver.get("http://" + ip_address)
        driver.maximize_window()

        # Fill in login form
        username_input = driver.find_element(By.NAME, "username")
        password_input = driver.find_element(By.NAME, "password")

        username_input.send_keys(username)
        password_input.send_keys(password)

        login_button = driver.find_element(By.NAME, "login")
        login_button.click()

        # Wait for the user to be logged in (adjust the time as needed)
        time.sleep(3)

        # Navigate to the data page
        driver.get(
            "http://103.139.165.110:8080/consumer/cons_order_reg.php")

        # Find the table element
        table_element = driver.find_element(
            By.CSS_SELECTOR, '.table')

        # Get the HTML content of the table
        table_html = table_element.get_attribute('outerHTML')

        # Use BeautifulSoup to parse the HTML
        soup = BeautifulSoup(table_html, 'html.parser')

        # Find all rows in the table body
        rows = soup.find_all('tbody')[0].find_all('tr')

        product_order_register = []

        # Loop through the rows and extract data
        for row in rows:
            cells = row.find_all('td')
            row_data = []  # List to store data for this row

            for cell in cells:
                cell_data = cell.text.strip()  # Get the text content of the cell
                row_data.append(cell_data)

        product_order_register.append(row_data)

    # Append scraped data to a shared result dictionary
    # Modify this according to your data
    scraped_data = {username: [product_order_register]}
    result_queue.put(scraped_data)
    print(scraped_data)
    # Close the browser window when done
    driver.quit()


def main():
    # Create a task queue
    task_queue = Queue()

    # Enqueue all tasks (username, password pairs) into the queue
    for username, password in zip(usernames, passwords):
        task_queue.put((username, password))

    # Create a result queue to store scraped data from processes
    result_queue = Queue()

    # Create and start concurrent processes
    processes = []
    for _ in range(MAX_CONCURRENT_PROCESSES):
        process = Process(target=scrape_data, args=(task_queue, result_queue))
        processes.append(process)
        process.start()

    # Wait for all processes to finish
    for process in processes:
        process.join()

    # Gather and process the scraped data from all processes
    all_data = {}
    while not result_queue.empty():
        scraped_data = result_queue.get()
        all_data.update(scraped_data)

    # Save all_data to a single JSON file
    with open("all_data.json", "w") as json_file:
        json.dump(all_data, json_file, indent=4)

        # write  in excell
    # Create a new Excel workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Write data to Excel
    for row_num, (key, value) in enumerate(all_data.items(), start=1):
        sheet.cell(row=row_num, column=1, value=key)
        sheet.cell(row=row_num, column=2, value=value)

    # Save the Excel file
    workbook.save("all_data.xlsx")


if __name__ == "__main__":
    main()
